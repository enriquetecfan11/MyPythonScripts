import pdfplumber
import pandas as pd
from openpyxl import Workbook

# Ruta del archivo PDF de entrada
pdf_path = "fitosanitariosOriginal.pdf"

# Crear un nuevo archivo de Excel
excel_path = "tabla_combinada.xlsx"
workbook = Workbook()
worksheet = workbook.active

# Abrir el archivo PDF
with pdfplumber.open(pdf_path) as pdf:
    row_offset = 0  # Desplazamiento de filas para cada tabla

    # Iterar sobre todas las páginas del PDF
    for i, page in enumerate(pdf.pages):
        # Extraer las tablas de la página
        tables = page.extract_tables()

        # Verificar si se encontraron tablas en la página actual
        if tables:
            # Iterar sobre las tablas de la página
            for table in tables:
                # Convertir la tabla en un DataFrame de Pandas
                df = pd.DataFrame(table[1:], columns=table[0])

                print(df.head())

                # Calcular la posición de inicio de la tabla en el archivo de Excel
                start_row = row_offset + 1
                start_column = 1

                # Escribir los encabezados en el archivo de Excel
                for j, header in enumerate(df.columns):
                    worksheet.cell(row=start_row, column=start_column + j, value=header)

                # Escribir los datos en el archivo de Excel
                for k, row in enumerate(df.values):
                    for j, value in enumerate(row):
                        worksheet.cell(row=start_row + k + 1, column=start_column + j, value=value)

                # Actualizar el desplazamiento de filas para la próxima tabla
                row_offset += len(df) + 2  # Añadir 2 filas de espacio entre tablas

    # Guardar el archivo de Excel
    workbook.save(excel_path)

    print(f"Archivo de Excel '{excel_path}' guardado con éxito.")
