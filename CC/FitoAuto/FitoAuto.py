import urllib.request
import PyPDF2
import pandas as pd
import numpy as np
import os
from datetime import date
from openpyxl import Workbook
import pdfplumber

# Paso 1: Descargar el archivo PDF
url = 'https://www.mapa.gob.es/es/agricultura/temas/sanidad-vegetal/listasustanciasactivasaceptadasexcluidas_tcm30-618972.pdf'
pdf_filename = 'archivo_{}.pdf'.format(date.today().strftime("%Y-%m-%d"))
urllib.request.urlretrieve(url, pdf_filename)

print("Paso 1 Ejecutado")

# Paso 2: Extraer páginas 1-48 del archivo PDF
extracted_pdf_filename = pdf_filename

with open(extracted_pdf_filename, 'rb') as file:
    pdf = PyPDF2.PdfReader(file)
    output_pdf = PyPDF2.PdfWriter()
    for page_num in range(0, 48):
        output_pdf.add_page(pdf.pages[page_num])
    with open(extracted_pdf_filename, 'wb') as extracted_pdf_file:
        output_pdf.write(extracted_pdf_file)

print("Paso 2 Ejecutado")


# Paso 3: Convertir el nuevo archivo PDF a XLSX
pdf_path = pdf_filename

# Crear un nuevo archivo de Excel
excel_path = "sustaciasFitosanitarios.xlsx"
workbook = Workbook()
worksheet = workbook.active

# Abrir el archivo PDF
with pdfplumber.open(pdf_path) as pdf:
    row_offset = 0  # Desplazamiento de filas para cada tabla

    # Iterar sobre todas las páginas del PDF
    for i, page in enumerate(pdf.pages):
        # Extraer las tablas de la página
        tables = page.extract_tables()

        # Verificar si se encontraron tablas en la página actual
        if tables:
            # Iterar sobre las tablas de la página
            for table in tables:
                # Convertir la tabla en un DataFrame de Pandas
                df = pd.DataFrame(table[1:], columns=table[0])

                # Calcular la posición de inicio de la tabla en el archivo de Excel
                start_row = row_offset + 1
                start_column = 1

                # Escribir los encabezados en el archivo de Excel
                for j, header in enumerate(df.columns):
                    worksheet.cell(row=start_row, column=start_column + j, value=header)

                # Escribir los datos en el archivo de Excel
                for k, row in enumerate(df.values):
                    for j, value in enumerate(row):
                        worksheet.cell(row=start_row + k + 1, column=start_column + j, value=value)

                # Actualizar el desplazamiento de filas para la próxima tabla
                row_offset += len(df) + 2  # Añadir 2 filas de espacio entre tablas

    # Guardar el archivo de Excel
    workbook.save(excel_path)
    print(f"Archivo de Excel '{excel_path}' guardado con éxito.")


print("Paso 3 Ejecutado")


# Paso 4: Modificaciones el exel
# Importas el archivo excel
excel_file = "sustaciasFitosanitarios.xlsx"

# Leer el archivo Excel en un DataFrame de pandas
df = pd.read_excel(excel_file)

# Imprimimos el numero de celdas del excel
print("Num celdas del excel: ")
print(df.shape)

# Sustimos los \n por un espacio
df = df.replace('\n', ' ', regex=True)

# Miramos si hay valores nulos y los sustituimos por espacios en blanco
df = df.replace(np.nan, ' ', regex=True)

# Miramos todas las celdas tengan espcaios en blanco y las eliminamos
df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

# Eliminar filas donde la primera columna comienza con un espacio en blanco
df = df[~df.iloc[:, 0].astype(str).str.startswith(' ')]

# Si la columa empieza con SUSTANCIA ACTIVA tambien se elimina
df = df[~df.iloc[:, 0].astype(str).str.startswith('SUSTANCIA ACTIVA')]

# Guardamos los cambios
df.to_excel("tabla_combinada_modificada.xlsx", index=False)
print(f"Archivo de Excel '{excel_file}' modificado con éxito.")

print("Paso 4 ejecutado")
